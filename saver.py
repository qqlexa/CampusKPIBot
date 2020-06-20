import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import os


class Saver:
    def __init__(self):
        self.working_dir = os.getcwd()

    def create_dir(self, path):
        if not os.path.exists(self.working_dir + path):
            os.makedirs(self.working_dir + path)

    def create_excel(self, path, speciality):
        if not os.path.exists(self.working_dir + path):
            writer = pd.ExcelWriter(self.working_dir + path, engine='openpyxl')

            pd.DataFrame().to_excel(writer, sheet_name=speciality, index=False)

            writer.save()

    @staticmethod
    def row_colour(row):
        return ['background-color:' + row.colour.lower() for i in row]

    def write_to_excel(self, info, average_mark, path, speciality, count_max):
        df_list = list()
        """
        count_max : marker green color
        """
        writer = pd.ExcelWriter(self.working_dir + path, engine='openpyxl')
        try:
            book = load_workbook(self.working_dir + path)
        except BaseException:
            pass
        else:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        if speciality in writer.book.sheetnames:
            startrow = writer.book[speciality].max_row

            df_list.append(pd.DataFrame([startrow + 1]))
            df_list.append(pd.DataFrame([info[0]]))
            df_list.append(pd.DataFrame([info[2]]))
            df_list.append(pd.DataFrame([average_mark]))

            for i in range(len(df_list)):
                df_list[i].to_excel(writer, header=None, sheet_name=speciality,
                                    startcol=i, startrow=startrow, index=False)
        else:
            print("Error with writing")
        sheet = writer.book.get_sheet_by_name(speciality)
        for col_range in range(1, count_max):
            for row_range in range(1, 4):
                cell_title = sheet.cell(row_range, col_range)
                cell_title.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        writer.save()

    def save_info(self, param, count_max):
        """

        :param param: list[info : list(), average_mark: int]
        :param count_max: int
        :return:
        """
        cols_names = ['Name', 'Faculty', 'Group', 'Studying form', 'Course', 'Speciality', 'Studying status']

        info = param[0]
        average_mark = param[1]

        dir_path = "/" + info[1] + "/" + info[4]
        excel_path = dir_path + "\\" + info[4] + ".xlsx"

        self.create_dir(dir_path)
        self.create_excel(excel_path, info[5])
        self.write_to_excel(info, average_mark, excel_path, info[5], int(count_max * 0.45))
